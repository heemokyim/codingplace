from openpyxl import load_workbook

wb = load_workbook('data/excel_writable.xlsx')

ws = wb['sheet_test']

# Modify certain cell of worksheet in excel
ws['A2'] = 'numb'

wb.save('data/excel_writable.xlsx')
