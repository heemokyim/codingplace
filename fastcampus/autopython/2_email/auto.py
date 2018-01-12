from auto_email import Email
from openpyxl import load_workbook

e = Email()

excel = load_workbook('email_address.xlsx',read_only=True)
worksheet = excel.active

for row in worksheet.iter_rows():
    name = row[0].value
    mail = row[1].value
    e.send_mail(name,mail)
