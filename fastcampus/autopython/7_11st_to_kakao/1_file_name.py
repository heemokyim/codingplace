from selenium import webdriver
from openpyxl import Workbook, load_workbook
from os.path import exists
from datetime import datetime
import platform
import string

today = datetime.now()
today_file_name = '%d_%d_%d.xlsx' % (today.year, today.month, today.day)

if exists(today_file_name):
    result_xlsx = load_workbook(today_file_name)

else:
    result_xlsx = Workbook()

worksheet = result_xlsx.active
worksheet['A2'] = '최근 가격'


try:
    driver = None
    if platform.system() == 'Windows':
        driver = './chromedriver.exe'
    else:
        driver = './chromedriver'

    driver = webdriver.Chrome(driver)
    driver.maximize_window()

    # 모든 상품 링크를 가져옴
    product_urls = open('products.txt','r').readlines()

    # strftime은 시간을 원하는 형태로 표기
    #                       시, 분, 초
    row = [today.strftime('%H:%M:%S')]

except Exception as e:
    print(e)

finally:
    driver.quit()
