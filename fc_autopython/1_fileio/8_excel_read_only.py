from openpyxl import load_workbook

# read_only=True -> Load data on the fly
excel_file = load_workbook('data/excel_writable.xlsx',read_only=True)

worksheet = excel_file['sheet_test']

for row in worksheet.iter_rows():
    for cell in row:
        print(cell.value)
