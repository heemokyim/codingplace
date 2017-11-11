from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet(title='sheet_test',index=1)
ws1 = wb.create_sheet(title='Sheet')

ws['A1'] = 'alghost'
ws['B1'] = 'test'

ws1['A1'] = 'leelee1234'

wb.save('data/excel_writable.xlsx')

print(wb.get_sheet_names())
print('Completed !')
