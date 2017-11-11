from openpyxl import load_workbook

wb = load_workbook('data/excel_writable.xlsx')

ws = wb['sheet_test']

ws['A2'] = 'numb'

wb.save('data/excel_writable.xlsx')
