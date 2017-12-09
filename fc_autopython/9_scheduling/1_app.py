from selenium import webdriver
from openpyxl import Workbook, load_workbook
from os.path import exists
from datetime import datetime
import string, os

# 이 소스가 속한 절대경로를 알려줌
base_dir = os.path.dirname(os.path.realpath(__file__))
print(base_dir)

today = datetime.now()
today_file_name = '%d_%d_%d.xlsx' % (today.year, today.month, today.day)

# 리눅스
file_path = '%s/%s' % (base_dir, today_file_name)

# 윈도우
# file_path = '%s\\%s' % (base_dir, today_file_name)

if exists(file_path):
    result_xlsx = load_workbook(file_path)
else:
    result_xlsx = Workbook()

worksheet = result_xlsx.active
worksheet['A2'] = '최근 가격'

opts = webdriver.ChromeOptions()
opts.add_argument('headless')
# 리눅스
driver = webdriver.Chrome(base_dir+'/chromedriver', chrome_options = opts)
# 윈도우
# driver = webdriver.Chrome(base_dir+'\\chromedriver', chrome_options = opts)
driver.maximize_window()

try:
    # 리눅스
    product_urls = open('products.txt','r').readlines()
    # 윈도우
    product_urls = open('products.txt','r').readlines()

    row = [today.strftime('%H:%M:%S')]

    changes=[]

    # 이 컬럼 인덱스 값으로 각 상품에 해당하는 컬럼(알파벳)으로 변환
    # 1 => B, 2 => C, 3 => D
    column_idx = 1

    # 저장된 링크를 돌면서 수행
    for product in product_urls:
        driver.get(product.strip())

        # div = driver.find_element_by_class_name('prdc_default_info')
        # price = div.find_element_by_tag_name('strong')
        # or
        # price_xpath = '//div[@contains(@class,"prdc_default_info")]/strong'
        # price = driver.find_element_by_xpath(price_xpath)

        # div = driver.find_element_by_class_name('prdc_default_info')
        # title = div.find_element_by_tag_name('h2')
        # or
        # title_xpath = '//div[@class='heading']/h2'
        # title = driver.find_element_by_xpath(title_xpath)

        price = driver.find_element_by_class_name('sale_price')
        row.append(price.text)

        print(driver.title)

        # 컬럼 인덱스를 컬럼 문자로 변경
        # ascii_uppercase는 'ABCDEF..'라는 문자열
        column = string.ascii_uppercase[column_idx]
        # 상품명을 갱신, 여기서 1은 행을 의미함 (결과는 B1,B2,B3,C1..)
        worksheet[column+'1'] = str(driver.title)

        prev_price = worksheet[column+'2'].value
        curr_price = price.text
        # 이전에 작성된 최근 가격과 비교하여
        if prev_price and prev_price != curr_price:
            # 다른 경우 tuple형태로 상품명, 이전가격, 최근가격 정보를 넣음
            print('Changes !')
            changes.append((driver.title, prev_price, curr_price))

        # 최근 가격을 갱신하고 컬럼 인덱스 증가
        worksheet[column+'2'] = price.text
        column_idx += 1

    # 시트에 행을 추가하고 액셀을 저장
    worksheet.append(row)
    result_xlsx.save(file_path)

    if changes:
        print(changes)
        from auto_email import Email
        contents = str(today)[:-7] + ' :가격이 변동된 상품이 있슴당.\n\n'

        for c in changes:
            contents += '(%s) %s => %s\n' % (c[0],c[1],c[2])

        e = Email()
        e.send_mail('이재웅','jkl2142@naver.com',contents=contents)

except Exception as e:
    print(e)
finally:
    driver.quit()
